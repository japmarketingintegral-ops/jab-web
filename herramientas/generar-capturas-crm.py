"""Dibuja las capturas del CRM para la nota, leyendo la planilla real.

No son mockups: los datos y los totales salen del mismo archivo que se entrega,
así que lo que se ve en la nota es exactamente lo que la persona se descarga.
"""
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from datetime import date

PLANILLA = '/Users/santiagociarniello/WEB/public/recursos/crm-para-vendedores-jab.xlsx'
SALIDA = '/Users/santiagociarniello/WEB/public/assets/img/blog'

NAVY = (0, 0, 46)
AZUL = (61, 106, 241)
BLANCO = (255, 255, 255)
TEXTO = (32, 36, 56)
SUAVE = (110, 116, 138)
LINEA = (223, 227, 238)
FONDO = (247, 248, 251)

ESTADOS = {
    'Atrasado':         ((255, 213, 213), (176, 0, 32)),
    'Es hoy':           ((255, 240, 194), (138, 97, 0)),
    'Sin próximo paso': ((255, 226, 194), (154, 75, 0)),
    'Al día':           ((220, 243, 227), (27, 110, 60)),
    'Ganado':           ((216, 230, 255), (27, 63, 160)),
    'Perdido':          ((237, 238, 242), (113, 118, 138)),
}

F = '/System/Library/Fonts/Avenir Next.ttc'
def fuente(px, negrita=False):
    return ImageFont.truetype(F, px, index=2 if negrita else 0)


def leer():
    wb = load_workbook(PLANILLA)
    op = wb['Oportunidades']
    hoy = date.today()
    filas = []
    for f in range(3, 203):
        emp = op.cell(f, 2).value
        if not emp:
            continue
        cuando = op.cell(f, 13).value
        cuando = cuando.date() if hasattr(cuando, 'date') else cuando
        ult = op.cell(f, 10).value
        ult = ult.date() if hasattr(ult, 'date') else ult
        etapa = op.cell(f, 9).value
        if etapa == 'Ganado':
            est = 'Ganado'
        elif etapa == 'Perdido':
            est = 'Perdido'
        elif not cuando:
            est = 'Sin próximo paso'
        elif cuando < hoy:
            est = 'Atrasado'
        elif cuando == hoy:
            est = 'Es hoy'
        else:
            est = 'Al día'
        filas.append({
            'empresa': emp, 'contacto': op.cell(f, 3).value, 'tel': op.cell(f, 4).value,
            'origen': op.cell(f, 6).value, 'valor': op.cell(f, 8).value or 0,
            'etapa': etapa, 'ultimo': ult, 'dias': (hoy - ult).days if ult else '',
            'paso': op.cell(f, 12).value or '', 'cuando': cuando, 'estado': est,
        })
    return filas


def pesos(n):
    return '$' + f'{int(n):,}'.replace(',', '.')


def chip(dr, x, y, texto, alto=22):
    fondo, tinta = ESTADOS[texto]
    f = fuente(12, True)
    ancho = dr.textlength(texto, font=f) + 20
    dr.rounded_rectangle([x, y, x + ancho, y + alto], radius=6, fill=fondo)
    dr.text((x + 10, y + alto / 2), texto, font=f, fill=tinta, anchor='lm')
    return ancho


def marco(ancho, alto, titulo, subtitulo):
    im = Image.new('RGB', (ancho, alto), BLANCO)
    dr = ImageDraw.Draw(im)
    dr.rectangle([0, 0, ancho, 54], fill=NAVY)
    dr.text((22, 27), titulo, font=fuente(17, True), fill=BLANCO, anchor='lm')
    dr.text((ancho - 22, 27), subtitulo, font=fuente(12), fill=(150, 165, 210), anchor='rm')
    return im, dr


# ------------------------------------------------------ 1 · La tabla completa
def tabla(filas, ruta, ancho=1000, alto=560):
    im, dr = marco(ancho, alto, 'OPORTUNIDADES', 'una fila por consulta')
    cols = [('Empresa', 22, 190), ('Contacto', 218, 120), ('Etapa', 348, 140),
            ('Valor', 496, 100), ('Días sin\nhablar', 610, 70), ('Próximo paso', 692, 170),
            ('Estado', 872, 110)]
    y = 54
    dr.rectangle([0, y, ancho, y + 40], fill=AZUL)
    for nombre, x, _ in cols:
        lineas = nombre.split('\n')
        for i, l in enumerate(lineas):
            dr.text((x, y + 20 + (i - (len(lineas) - 1) / 2) * 12), l,
                    font=fuente(11, True), fill=BLANCO, anchor='lm')
    y += 40

    fila_h = 38
    for r in filas[:12]:
        if y + fila_h > alto:
            break
        if r['estado'] in ('Atrasado', 'Es hoy'):
            dr.rectangle([0, y, ancho, y + fila_h], fill=(253, 250, 246))
        dr.line([0, y + fila_h, ancho, y + fila_h], fill=LINEA)
        cy = y + fila_h / 2
        dr.text((cols[0][1], cy), r['empresa'][:24], font=fuente(12, True), fill=TEXTO, anchor='lm')
        dr.text((cols[1][1], cy), r['contacto'][:16], font=fuente(12), fill=SUAVE, anchor='lm')
        dr.text((cols[2][1], cy), r['etapa'], font=fuente(12), fill=TEXTO, anchor='lm')
        dr.text((cols[3][1] + 80, cy), pesos(r['valor']), font=fuente(12), fill=TEXTO, anchor='rm')
        d = r['dias']
        if d != '' and r['etapa'] not in ('Ganado', 'Perdido'):
            dr.text((cols[4][1] + 30, cy), str(d), font=fuente(12, True),
                    fill=(176, 0, 32) if d > 7 else TEXTO, anchor='mm')
        paso = r['paso'][:24] + ('…' if len(r['paso']) > 24 else '') if r['paso'] else '—'
        dr.text((cols[5][1], cy), paso, font=fuente(11), fill=SUAVE if r['paso'] else (200, 120, 60), anchor='lm')
        chip(dr, cols[6][1], y + 8, r['estado'])
        y += fila_h
    im.save(ruta, quality=92)


# ------------------------------------------------------------- 2 · La hoja Hoy
def hoja_hoy(filas, ruta, ancho=1000, alto=560):
    im, dr = marco(ancho, alto, 'HOY', 'a quién tenés que llamar')
    pend = [r for r in filas if r['estado'] in ('Atrasado', 'Es hoy', 'Sin próximo paso')]
    orden = {'Atrasado': 0, 'Es hoy': 1, 'Sin próximo paso': 2}
    pend.sort(key=lambda r: orden[r['estado']])

    dr.text((22, 82), 'Se llena sola con lo que cargaste. Esto es tu día:',
            font=fuente(14), fill=SUAVE, anchor='lm')

    y = 112
    for r in pend:
        alto_f = 74
        dr.rounded_rectangle([22, y, ancho - 22, y + alto_f], radius=10, fill=FONDO)
        dr.rectangle([22, y, 27, y + alto_f], fill=ESTADOS[r['estado']][1])
        chip(dr, 42, y + 13, r['estado'], alto=20)
        dr.text((ancho - 42, y + 23), r['tel'], font=fuente(13), fill=SUAVE, anchor='rm')
        dr.text((42, y + 50), r['empresa'], font=fuente(15, True), fill=TEXTO, anchor='lm')
        w = dr.textlength(r['empresa'], font=fuente(15, True))
        dr.text((42 + w + 14, y + 51), '·  ' + (r['paso'] or 'definí el próximo paso'),
                font=fuente(13), fill=SUAVE if r['paso'] else (154, 75, 0), anchor='lm')
        y += alto_f + 10
        if y + 74 > alto:
            break
    im.save(ruta, quality=92)


# --------------------------------------------------------------- 3 · Tablero
def tablero(filas, ruta, ancho=1000, alto=440):
    im, dr = marco(ancho, alto, 'TABLERO', 'se calcula solo')
    ab = [r for r in filas if r['etapa'] not in ('Ganado', 'Perdido')]
    ga = [r for r in filas if r['etapa'] == 'Ganado']
    pe = [r for r in filas if r['etapa'] == 'Perdido']

    tarjetas = [
        ('Oportunidades abiertas', str(len(ab)), TEXTO),
        ('Plata en juego', pesos(sum(r['valor'] for r in ab)), AZUL),
        ('Atrasadas', str(sum(1 for r in filas if r['estado'] == 'Atrasado')), (176, 0, 32)),
        ('Sin próximo paso', str(sum(1 for r in filas if r['estado'] == 'Sin próximo paso')), (154, 75, 0)),
    ]
    x, y, w, h = 22, 82, (ancho - 44 - 3 * 12) / 4, 96
    for nombre, valor, color in tarjetas:
        dr.rounded_rectangle([x, y, x + w, y + h], radius=10, fill=FONDO)
        dr.text((x + 16, y + 26), nombre, font=fuente(11), fill=SUAVE, anchor='lm')
        dr.text((x + 16, y + 62), valor, font=fuente(26 if len(valor) < 8 else 19, True),
                fill=color, anchor='lm')
        x += w + 12

    y += h + 26
    dr.text((22, y), 'Dónde está cada oportunidad', font=fuente(14, True), fill=NAVY, anchor='lm')
    y += 22
    etapas = ['Nuevo', 'Contactado', 'Reunión hecha', 'Presupuesto enviado', 'Negociando']
    tope = max(sum(r['valor'] for r in filas if r['etapa'] == e) for e in etapas) or 1
    for e in etapas:
        g = [r for r in filas if r['etapa'] == e]
        v = sum(r['valor'] for r in g)
        dr.text((22, y + 15), e, font=fuente(12), fill=TEXTO, anchor='lm')
        bx = 190
        largo = (ancho - 22 - bx - 130) * (v / tope)
        dr.rounded_rectangle([bx, y + 6, bx + max(largo, 3), y + 24], radius=4, fill=AZUL)
        dr.text((bx + max(largo, 3) + 10, y + 15), f'{len(g)}   {pesos(v)}',
                font=fuente(12, True), fill=TEXTO, anchor='lm')
        y += 30

    y += 8
    dr.line([22, y, ancho - 22, y], fill=LINEA)
    y += 20
    cerradas = len(ga) + len(pe)
    resumen = f'Cerradas: {len(ga)} ganadas · {len(pe)} perdidas   ·   Facturado {pesos(sum(r["valor"] for r in ga))}'
    if cerradas:
        resumen += f'   ·   De cada 10 que cerrás, ganás {round(len(ga)/cerradas*10, 1)}'
    dr.text((22, y), resumen, font=fuente(13), fill=SUAVE, anchor='lm')
    im.save(ruta, quality=92)


# ----------------------------------------------------------------- 4 · Portada
def portada(filas, ruta, ancho=1200, alto=675):
    im = Image.new('RGB', (ancho, alto), NAVY)
    dr = ImageDraw.Draw(im)
    dr.text((70, 128), 'CRM PARA VENDEDORES', font=fuente(15, True), fill=AZUL)
    dr.text((70, 168), 'Nunca más', font=fuente(58, True), fill=(224, 244, 251))
    dr.text((70, 234), 'una consulta perdida', font=fuente(58, True), fill=(224, 244, 251))
    dr.text((70, 322), 'Una planilla gratis para que sepas, cada mañana,', font=fuente(19), fill=(150, 165, 210))
    dr.text((70, 352), 'a quién tenés que llamar.', font=fuente(19), fill=(150, 165, 210))

    # Un recorte de la hoja Hoy, como si estuviera apoyado
    caja_y = 430
    dr.rounded_rectangle([70, caja_y, ancho - 70, alto - 46], radius=14, fill=BLANCO)
    pend = [r for r in filas if r['estado'] in ('Atrasado', 'Es hoy', 'Sin próximo paso')][:2]
    yy = caja_y + 22
    for r in pend:
        chip(dr, 96, yy, r['estado'], alto=22)
        dr.text((96, yy + 46), r['empresa'], font=fuente(16, True), fill=TEXTO, anchor='lm')
        w = dr.textlength(r['empresa'], font=fuente(16, True))
        dr.text((96 + w + 16, yy + 47), '·  ' + (r['paso'] or 'definí el próximo paso'),
                font=fuente(14), fill=SUAVE, anchor='lm')
        dr.text((ancho - 96, yy + 11), r['tel'], font=fuente(13), fill=SUAVE, anchor='rm')
        yy += 82
    im.save(ruta, quality=92)


filas = leer()
tabla(filas, f'{SALIDA}/crm-gratis-para-vendedores-2.jpg')
hoja_hoy(filas, f'{SALIDA}/crm-gratis-para-vendedores-3.jpg')
tablero(filas, f'{SALIDA}/crm-gratis-para-vendedores-4.jpg')
portada(filas, f'{SALIDA}/crm-gratis-para-vendedores.jpg')
print('4 imágenes generadas desde', len(filas), 'filas reales')
