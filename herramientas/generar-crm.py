"""Arma el CRM en planilla que se entrega como recurso gratuito.

Pensado para un vendedor que hoy trabaja con la libreta y el WhatsApp: una sola
tabla de oportunidades, alertas automáticas de a quién hay que llamar, y un
tablero que se llena solo. Se sube a Google Sheets tal cual.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

NAVY = '00002E'
AZUL = '3D6AF1'
CELESTE = 'E0F4FB'
GRIS = 'F2F4F8'
BLANCO = 'FFFFFF'
AMARILLO = 'FFF3C4'

FILAS = 200          # hasta dónde llegan las fórmulas preparadas
PRIMERA = 3          # primera fila de datos

wb = Workbook()

titulo_f = Font(name='Arial', size=16, bold=True, color=BLANCO)
cab_f = Font(name='Arial', size=10, bold=True, color=BLANCO)
normal_f = Font(name='Arial', size=10)
chico_f = Font(name='Arial', size=9, color='5A5F73')
negrita_f = Font(name='Arial', size=10, bold=True)

navy_fill = PatternFill('solid', fgColor=NAVY)
azul_fill = PatternFill('solid', fgColor=AZUL)
gris_fill = PatternFill('solid', fgColor=GRIS)
amarillo_fill = PatternFill('solid', fgColor=AMARILLO)

borde_fino = Border(bottom=Side(style='thin', color='D5DAE5'))


# ---------------------------------------------------------------- Empezá acá
ini = wb.active
ini.title = 'Empezá acá'
ini.sheet_view.showGridLines = False
ini.column_dimensions['A'].width = 3
ini.column_dimensions['B'].width = 105

ini['B2'] = 'CRM para vendedores'
ini['B2'].font = Font(name='Arial', size=22, bold=True, color=NAVY)
ini['B3'] = 'Hecho por Jab Marketing  ·  jabmarketing.site'
ini['B3'].font = Font(name='Arial', size=11, color=AZUL)

bloques = [
    ('Qué es esto', [
        'Una planilla para que ninguna consulta se te pierda. No reemplaza a un CRM de verdad,',
        'pero hace lo que importa: te dice a quién tenés que llamar hoy y qué pasó con cada consulta.',
        'Si llega un día en que se te queda corta, es buena señal: quiere decir que estás vendiendo más.',
    ]),
    ('Cómo se usa, en tres pasos', [
        '1.  Cada consulta que entra, la cargás en la hoja OPORTUNIDADES. Una fila por consulta.',
        '2.  Cada vez que hablás con alguien, actualizás dos celdas: "Último contacto" y "Próximo paso".',
        '3.  A la mañana abrís la hoja HOY y llamás a los que aparecen ahí. Eso es todo.',
    ]),
    ('Las únicas dos celdas que no podés dejar vacías', [
        'PRÓXIMO PASO y CUÁNDO. Sin eso, la consulta queda a la deriva y la planilla te lo va a marcar.',
        'Aunque sea "llamar para ver si lo pensó" dentro de diez días: siempre tiene que haber un próximo paso.',
    ]),
    ('Qué significan los colores', [
        'Atrasado          la fecha del próximo paso ya pasó. Es lo primero que hay que resolver.',
        'Es hoy            tenés que hacer algo con este contacto en el día.',
        'Sin próximo paso  cargaste la consulta pero no definiste qué sigue. Se pierde así.',
        'Al día            no hay que hacer nada todavía.',
    ]),
    ('Antes de arrancar', [
        'La hoja OPORTUNIDADES viene con doce filas de ejemplo para que veas cómo se completa',
        'y para que el tablero muestre algo. Borralas y empezá con las tuyas.',
        'En la hoja LISTAS podés cambiar las etapas y los orígenes por los que uses vos.',
    ]),
]

fila = 5
for titulo, lineas in bloques:
    ini.cell(fila, 2, titulo).font = Font(name='Arial', size=12, bold=True, color=NAVY)
    ini.cell(fila, 2).fill = gris_fill
    fila += 1
    for l in lineas:
        c = ini.cell(fila, 2, l)
        c.font = normal_f
        c.alignment = Alignment(vertical='center')
        fila += 1
    fila += 1

ini.cell(fila, 2, '¿Se te quedó corta, o querés que el seguimiento se haga solo?').font = negrita_f
ini.cell(fila + 1, 2, 'Escribinos a jabmarketing.site — implementamos CRM de verdad, con el WhatsApp integrado.').font = normal_f


# ------------------------------------------------------------------- Listas
lis = wb.create_sheet('Listas')
lis.sheet_view.showGridLines = False
lis['A1'] = 'ETAPAS'
lis['B1'] = 'DE DÓNDE SALIÓ'
for c in ('A1', 'B1'):
    lis[c].font = cab_f
    lis[c].fill = navy_fill
lis.column_dimensions['A'].width = 24
lis.column_dimensions['B'].width = 24

ETAPAS = ['Nuevo', 'Contactado', 'Reunión hecha', 'Presupuesto enviado',
          'Negociando', 'Ganado', 'Perdido']
ORIGENES = ['Referido', 'WhatsApp', 'Instagram', 'Facebook', 'Google',
            'Llamada en frío', 'Vino al local', 'Evento / feria', 'Cliente anterior', 'Otro']

for i, e in enumerate(ETAPAS, start=2):
    lis.cell(i, 1, e).font = normal_f
for i, o in enumerate(ORIGENES, start=2):
    lis.cell(i, 2, o).font = normal_f

lis['D1'] = 'Cambiá estas listas por las tuyas y los desplegables se actualizan solos.'
lis['D1'].font = chico_f
lis['D2'] = 'Ganado y Perdido tienen que quedar con ese nombre: el tablero los busca así.'
lis['D2'].font = chico_f


# ------------------------------------------------------------ Oportunidades
op = wb.create_sheet('Oportunidades', 1)
op.sheet_view.showGridLines = False

COLS = [
    ('Fecha de alta', 13, 'DD/MM/YYYY'),
    ('Empresa', 24, None),
    ('Contacto', 20, None),
    ('Teléfono', 16, None),
    ('Mail', 26, None),
    ('De dónde salió', 17, None),
    ('Qué necesita', 34, None),
    ('Valor estimado', 15, '"$"#,##0'),
    ('Etapa', 20, None),
    ('Último contacto', 15, 'DD/MM/YYYY'),
    ('Días sin hablar', 14, '0'),
    ('Próximo paso', 30, None),
    ('Cuándo', 13, 'DD/MM/YYYY'),
    ('Estado', 18, None),
]

op.merge_cells('A1:N1')
op['A1'] = 'OPORTUNIDADES     ·     una fila por consulta'
op['A1'].font = titulo_f
op['A1'].fill = navy_fill
op['A1'].alignment = Alignment(vertical='center', indent=1)
op.row_dimensions[1].height = 30

for i, (nombre, ancho, fmt) in enumerate(COLS, start=1):
    c = op.cell(2, i, nombre)
    c.font = cab_f
    c.fill = azul_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    op.column_dimensions[get_column_letter(i)].width = ancho
op.row_dimensions[2].height = 30

hoy = date.today()
EJEMPLOS = [
    (hoy - timedelta(days=2),  'Metalúrgica San Jorge', 'Roberto Díaz', '341 555-0182', 'rdiaz@sanjorge.com.ar', 'Referido', 'Necesita 200 piezas por mes, hoy le compra a Buenos Aires', 850000, 'Presupuesto enviado', hoy - timedelta(days=1), 'Llamar para saber si lo revisó', hoy + timedelta(days=2)),
    (hoy - timedelta(days=9),  'Corralón El Puente', 'Silvia Gómez', '341 555-0143', 'ventas@elpuente.com.ar', 'WhatsApp', 'Pidió lista de precios mayorista', 420000, 'Negociando', hoy - timedelta(days=3), 'Mandar la contrapropuesta', hoy - timedelta(days=1)),
    (hoy - timedelta(days=15), 'Transporte Fighiera', 'Luis Ferrero', '3402 55-0119', 'luis@fighiera.com.ar', 'Referido', 'Quiere renovar toda la flota de cubiertas', 1250000, 'Reunión hecha', hoy - timedelta(days=6), 'Pasar el presupuesto final', hoy),
    (hoy - timedelta(days=1),  'Panadería Doña Elsa', 'Elsa Martínez', '341 555-0167', '', 'Instagram', 'Consultó por la máquina chica', 180000, 'Nuevo', hoy - timedelta(days=1), 'Llamarla y entender qué volumen maneja', hoy + timedelta(days=1)),
    (hoy - timedelta(days=22), 'Agropecuaria del Sur', 'Marcos Rivas', '3402 55-0188', 'mrivas@agrosur.com.ar', 'Evento / feria', 'Lo conocimos en Expoagro, maneja 400 hectáreas', 2100000, 'Negociando', hoy - timedelta(days=11), 'Insistir, quedó en confirmar', hoy - timedelta(days=4)),
    (hoy - timedelta(days=30), 'Ferretería Central', 'Ana Suárez', '341 555-0155', 'anasuarez@fcentral.com', 'Vino al local', 'Compra chica pero repite todos los meses', 95000, 'Ganado', hoy - timedelta(days=25), '', None),
    (hoy - timedelta(days=41), 'Constructora Belgrano', 'Pablo Ojeda', '341 555-0134', 'pojeda@belgrano.com.ar', 'Google', 'Obra grande, pedía financiación a 90 días', 3400000, 'Perdido', hoy - timedelta(days=33), '', None),
    (hoy - timedelta(days=5),  'Distribuidora Norte', 'Carla Benítez', '341 555-0171', 'cbenitez@dnorte.com.ar', 'Referido', 'Le interesó el combo de temporada', 660000, 'Contactado', hoy - timedelta(days=5), 'Mandarle el catálogo', hoy + timedelta(days=3)),
    (hoy - timedelta(days=12), 'Taller Rodríguez', 'Juan Rodríguez', '3402 55-0142', '', 'Llamada en frío', 'Atendió pero estaba ocupado, pidió que lo llame la semana que viene', 240000, 'Contactado', hoy - timedelta(days=12), '', None),
    (hoy - timedelta(days=3),  'Vivero Los Álamos', 'Marta Kiener', '341 555-0129', 'info@losalamos.com.ar', 'Facebook', 'Quiere el sistema de riego completo', 540000, 'Reunión hecha', hoy - timedelta(days=2), 'Enviar propuesta con dos opciones', hoy + timedelta(days=1)),
    (hoy - timedelta(days=55), 'Frigorífico del Litoral', 'Sergio Paz', '341 555-0198', 'spaz@litoral.com.ar', 'Cliente anterior', 'Ya nos compró en 2024, quiere ampliar', 1800000, 'Ganado', hoy - timedelta(days=48), '', None),
    (hoy - timedelta(days=7),  'Almacén Doña Rosa', 'Rosa Aguirre', '341 555-0113', '', 'WhatsApp', 'Preguntó precio y no contestó más', 75000, 'Nuevo', hoy - timedelta(days=7), '', None),
]

# Las columnas 11 (días sin hablar) y 14 (estado) son fórmulas: los datos de
# ejemplo saltean la 11 y terminan en la 13.
DESTINO = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13]
for n, datos in enumerate(EJEMPLOS):
    f = PRIMERA + n
    assert len(datos) == len(DESTINO), f'fila {n}: {len(datos)} valores'
    for col, v in zip(DESTINO, datos):
        c = op.cell(f, col, v)
        c.font = normal_f

# Fórmulas y formato en todas las filas preparadas
for f in range(PRIMERA, PRIMERA + FILAS):
    # Días sin hablar: sólo mientras la oportunidad sigue viva
    op.cell(f, 11).value = (
        f'=IF($B{f}="","",'
        f'IF(OR($I{f}="Ganado",$I{f}="Perdido"),"",'
        f'IF($J{f}="","",TODAY()-$J{f})))'
    )
    # Estado: la alerta que ordena el día
    op.cell(f, 14).value = (
        f'=IF($B{f}="","",'
        f'IF($I{f}="Ganado","Ganado",'
        f'IF($I{f}="Perdido","Perdido",'
        f'IF($M{f}="","Sin próximo paso",'
        f'IF($M{f}<TODAY(),"Atrasado",'
        f'IF($M{f}=TODAY(),"Es hoy","Al día"))))))'
    )
    for i in range(1, 15):
        c = op.cell(f, i)
        if c.font.name != 'Arial':
            c.font = normal_f
        c.border = borde_fino
        _, _, fmt = COLS[i - 1]
        if fmt:
            c.number_format = fmt
    op.cell(f, 11).font = negrita_f
    op.cell(f, 11).alignment = Alignment(horizontal='center')
    op.cell(f, 14).font = negrita_f
    op.cell(f, 14).alignment = Alignment(horizontal='center')
    # Las celdas que el vendedor completa a mano quedan resaltadas
    op.cell(f, 12).fill = amarillo_fill
    op.cell(f, 13).fill = amarillo_fill

rango = f'A{PRIMERA}:N{PRIMERA + FILAS - 1}'

dv_etapa = DataValidation(type='list', formula1=f'Listas!$A$2:$A${len(ETAPAS) + 1}', allow_blank=True)
dv_origen = DataValidation(type='list', formula1=f'Listas!$B$2:$B${len(ORIGENES) + 1}', allow_blank=True)
op.add_data_validation(dv_etapa)
op.add_data_validation(dv_origen)
dv_etapa.add(f'I{PRIMERA}:I{PRIMERA + FILAS - 1}')
dv_origen.add(f'F{PRIMERA}:F{PRIMERA + FILAS - 1}')

est = f'N{PRIMERA}:N{PRIMERA + FILAS - 1}'
op.conditional_formatting.add(est, CellIsRule(operator='equal', formula=['"Atrasado"'],
    fill=PatternFill('solid', fgColor='FFD5D5'), font=Font(name='Arial', size=10, bold=True, color='B00020')))
op.conditional_formatting.add(est, CellIsRule(operator='equal', formula=['"Es hoy"'],
    fill=PatternFill('solid', fgColor='FFF0C2'), font=Font(name='Arial', size=10, bold=True, color='8A6100')))
op.conditional_formatting.add(est, CellIsRule(operator='equal', formula=['"Sin próximo paso"'],
    fill=PatternFill('solid', fgColor='FFE2C2'), font=Font(name='Arial', size=10, bold=True, color='9A4B00')))
op.conditional_formatting.add(est, CellIsRule(operator='equal', formula=['"Al día"'],
    fill=PatternFill('solid', fgColor='DCF3E3'), font=Font(name='Arial', size=10, color='1B6E3C')))
op.conditional_formatting.add(est, CellIsRule(operator='equal', formula=['"Ganado"'],
    fill=PatternFill('solid', fgColor='D8E6FF'), font=Font(name='Arial', size=10, bold=True, color='1B3FA0')))
op.conditional_formatting.add(est, CellIsRule(operator='equal', formula=['"Perdido"'],
    fill=PatternFill('solid', fgColor='EDEEF2'), font=Font(name='Arial', size=10, color='71768A')))

dias = f'K{PRIMERA}:K{PRIMERA + FILAS - 1}'
op.conditional_formatting.add(dias, CellIsRule(operator='greaterThan', formula=['7'],
    font=Font(name='Arial', size=10, bold=True, color='B00020')))

op.freeze_panes = 'C3'
op.auto_filter.ref = f'A2:N{PRIMERA + FILAS - 1}'


# ---------------------------------------------------------------------- Hoy
hoja_hoy = wb.create_sheet('Hoy', 2)
hoja_hoy.sheet_view.showGridLines = False
hoja_hoy.merge_cells('A1:F1')
hoja_hoy['A1'] = 'HOY     ·     a quién tenés que llamar'
hoja_hoy['A1'].font = titulo_f
hoja_hoy['A1'].fill = navy_fill
hoja_hoy['A1'].alignment = Alignment(vertical='center', indent=1)
hoja_hoy.row_dimensions[1].height = 30

hoja_hoy['A3'] = ('Esta hoja se llena sola con lo que cargás en OPORTUNIDADES. '
                  'Filtrá la columna Estado por Atrasado, Es hoy y Sin próximo paso: eso es tu día.')
hoja_hoy['A3'].font = chico_f

CAB_HOY = [('Estado', 18), ('Empresa', 24), ('Contacto', 20), ('Teléfono', 16),
           ('Próximo paso', 34), ('Cuándo', 13)]
for i, (n, w) in enumerate(CAB_HOY, start=1):
    c = hoja_hoy.cell(5, i, n)
    c.font = cab_f
    c.fill = azul_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    hoja_hoy.column_dimensions[get_column_letter(i)].width = w
hoja_hoy.row_dimensions[5].height = 24

# Espejo de la tabla: se trae todo y se filtra a ojo, que es lo que se puede
# hacer sin funciones modernas y funciona igual en Excel y en Google Sheets.
ORIGEN = {1: 'N', 2: 'B', 3: 'C', 4: 'D', 5: 'L', 6: 'M'}
for n in range(FILAS):
    f_dest = 6 + n
    f_orig = PRIMERA + n
    for col, letra in ORIGEN.items():
        c = hoja_hoy.cell(f_dest, col)
        c.value = (f'=IF(Oportunidades!$B{f_orig}="","",'
                   f'IF(OR(Oportunidades!$N{f_orig}="Ganado",Oportunidades!$N{f_orig}="Perdido",'
                   f'Oportunidades!$N{f_orig}="Al día"),"",Oportunidades!${letra}{f_orig}))')
        c.font = normal_f
        c.border = borde_fino
    hoja_hoy.cell(f_dest, 1).font = negrita_f
    hoja_hoy.cell(f_dest, 1).alignment = Alignment(horizontal='center')
    hoja_hoy.cell(f_dest, 6).number_format = 'DD/MM/YYYY'

rango_hoy = f'A6:A{5 + FILAS}'
hoja_hoy.conditional_formatting.add(rango_hoy, CellIsRule(operator='equal', formula=['"Atrasado"'],
    fill=PatternFill('solid', fgColor='FFD5D5'), font=Font(name='Arial', size=10, bold=True, color='B00020')))
hoja_hoy.conditional_formatting.add(rango_hoy, CellIsRule(operator='equal', formula=['"Es hoy"'],
    fill=PatternFill('solid', fgColor='FFF0C2'), font=Font(name='Arial', size=10, bold=True, color='8A6100')))
hoja_hoy.conditional_formatting.add(rango_hoy, CellIsRule(operator='equal', formula=['"Sin próximo paso"'],
    fill=PatternFill('solid', fgColor='FFE2C2'), font=Font(name='Arial', size=10, bold=True, color='9A4B00')))
hoja_hoy.freeze_panes = 'A6'
hoja_hoy.auto_filter.ref = f'A5:F{5 + FILAS}'


# ------------------------------------------------------------------ Tablero
tab = wb.create_sheet('Tablero', 3)
tab.sheet_view.showGridLines = False
tab.merge_cells('A1:F1')
tab['A1'] = 'TABLERO     ·     se calcula solo'
tab['A1'].font = titulo_f
tab['A1'].fill = navy_fill
tab['A1'].alignment = Alignment(vertical='center', indent=1)
tab.row_dimensions[1].height = 30

for col, w in zip('ABCDEF', (30, 18, 6, 30, 18, 18)):
    tab.column_dimensions[col].width = w

R = f'$I${PRIMERA}:$I${PRIMERA + FILAS - 1}'          # etapa
V = f'$H${PRIMERA}:$H${PRIMERA + FILAS - 1}'          # valor
B = f'$B${PRIMERA}:$B${PRIMERA + FILAS - 1}'          # empresa
N = f'$N${PRIMERA}:$N${PRIMERA + FILAS - 1}'          # estado

def bloque(fila, titulo):
    tab.cell(fila, 1, titulo).font = Font(name='Arial', size=12, bold=True, color=NAVY)
    tab.cell(fila, 1).fill = gris_fill
    tab.cell(fila, 2).fill = gris_fill

bloque(3, 'Tu embudo hoy')
kpis = [
    ('Oportunidades abiertas', f'=COUNTIFS(Oportunidades!{B},"<>",Oportunidades!{R},"<>Ganado",Oportunidades!{R},"<>Perdido")', '0'),
    ('Plata en juego', f'=SUMIFS(Oportunidades!{V},Oportunidades!{B},"<>",Oportunidades!{R},"<>Ganado",Oportunidades!{R},"<>Perdido")', '"$"#,##0'),
    ('Atrasadas', f'=COUNTIF(Oportunidades!{N},"Atrasado")', '0'),
    ('Sin próximo paso', f'=COUNTIF(Oportunidades!{N},"Sin próximo paso")', '0'),
    ('Para hacer hoy', f'=COUNTIF(Oportunidades!{N},"Es hoy")', '0'),
]
for n, (nombre, formula, fmt) in enumerate(kpis):
    f = 4 + n
    tab.cell(f, 1, nombre).font = normal_f
    c = tab.cell(f, 2, formula)
    c.font = Font(name='Arial', size=12, bold=True, color=NAVY)
    c.number_format = fmt
    c.alignment = Alignment(horizontal='right')
    tab.cell(f, 1).border = borde_fino
    tab.cell(f, 2).border = borde_fino

bloque(10, 'Cómo venís cerrando')
cierre = [
    ('Ganadas', f'=COUNTIF(Oportunidades!{R},"Ganado")', '0'),
    ('Perdidas', f'=COUNTIF(Oportunidades!{R},"Perdido")', '0'),
    ('Facturado', f'=SUMIF(Oportunidades!{R},"Ganado",Oportunidades!{V})', '"$"#,##0'),
    ('De cada 10 cerradas, ganás', f'=IFERROR(ROUND(COUNTIF(Oportunidades!{R},"Ganado")/(COUNTIF(Oportunidades!{R},"Ganado")+COUNTIF(Oportunidades!{R},"Perdido"))*10,1),0)', '0.0'),
    ('Valor promedio de una venta', f'=IFERROR(SUMIF(Oportunidades!{R},"Ganado",Oportunidades!{V})/COUNTIF(Oportunidades!{R},"Ganado"),0)', '"$"#,##0'),
]
for n, (nombre, formula, fmt) in enumerate(cierre):
    f = 11 + n
    tab.cell(f, 1, nombre).font = normal_f
    c = tab.cell(f, 2, formula)
    c.font = Font(name='Arial', size=12, bold=True, color=NAVY)
    c.number_format = fmt
    c.alignment = Alignment(horizontal='right')
    tab.cell(f, 1).border = borde_fino
    tab.cell(f, 2).border = borde_fino

# Embudo por etapa
tab.cell(3, 4, 'Dónde está cada oportunidad').font = Font(name='Arial', size=12, bold=True, color=NAVY)
for col in (4, 5, 6):
    tab.cell(3, col).fill = gris_fill
for i, n in enumerate(('Etapa', 'Cuántas', 'Cuánta plata')):
    c = tab.cell(4, 4 + i, n)
    c.font = cab_f
    c.fill = azul_fill
    c.alignment = Alignment(horizontal='center')

for n, etapa in enumerate(ETAPAS):
    f = 5 + n
    tab.cell(f, 4, etapa).font = normal_f
    c1 = tab.cell(f, 5, f'=COUNTIF(Oportunidades!{R},$D{f})')
    c2 = tab.cell(f, 6, f'=SUMIF(Oportunidades!{R},$D{f},Oportunidades!{V})')
    c1.number_format = '0'
    c2.number_format = '"$"#,##0'
    for c in (c1, c2):
        c.font = negrita_f
        c.alignment = Alignment(horizontal='right')
    for col in (4, 5, 6):
        tab.cell(f, col).border = borde_fino

tab.cell(14, 4, 'De dónde salen tus consultas').font = Font(name='Arial', size=12, bold=True, color=NAVY)
for col in (4, 5, 6):
    tab.cell(14, col).fill = gris_fill
for i, n in enumerate(('Origen', 'Consultas', 'Ganadas')):
    c = tab.cell(15, 4 + i, n)
    c.font = cab_f
    c.fill = azul_fill
    c.alignment = Alignment(horizontal='center')

O = f'$F${PRIMERA}:$F${PRIMERA + FILAS - 1}'
for n, origen in enumerate(ORIGENES):
    f = 16 + n
    tab.cell(f, 4, origen).font = normal_f
    c1 = tab.cell(f, 5, f'=COUNTIF(Oportunidades!{O},$D{f})')
    c2 = tab.cell(f, 6, f'=COUNTIFS(Oportunidades!{O},$D{f},Oportunidades!{R},"Ganado")')
    for c in (c1, c2):
        c.number_format = '0'
        c.font = negrita_f
        c.alignment = Alignment(horizontal='right')
    for col in (4, 5, 6):
        tab.cell(f, col).border = borde_fino

tab.cell(27, 4, 'Esta última tabla es la que te dice dónde poner el esfuerzo:').font = chico_f
tab.cell(28, 4, 'el origen que más consultas trae no siempre es el que más ventas cierra.').font = chico_f

tab.cell(17, 1, 'Todo se calcula solo desde OPORTUNIDADES.').font = chico_f
tab.cell(18, 1, 'No escribas nada en esta hoja.').font = chico_f

wb.save('/Users/santiagociarniello/WEB/public/recursos/crm-para-vendedores-jab.xlsx')
print('guardado')
